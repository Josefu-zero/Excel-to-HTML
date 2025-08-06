
## Author: Joseph A. Jimenez J.
## Josefu-Zero

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import os
import requests
from io import BytesIO
import re
from dotenv import load_dotenv
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.runtime.client_request_exception import ClientRequestException
from office365.sharepoint.client_context import ClientContext

# Cargar variables del entorno
load_dotenv()  # Busca automáticamente el archivo .env


def listar_archivos_excel_en_sharepoint():
    try:
        sitio_url = os.getenv("SHAREPOINT_URL")
        sitio_completo = f"{sitio_url}/{os.getenv('SHAREPOINT_SITE')}"
        ctx_auth = AuthenticationContext(sitio_completo)
        print(f"Autenticando en {sitio_completo} con usuario {os.getenv('SHAREPOINT_USER')}")
        if not ctx_auth.acquire_token_for_user(os.getenv("SHAREPOINT_USER"), os.getenv("SHAREPOINT_PASSWORD")):
            raise ValueError("Error de autenticación")

        ctx = ClientContext(sitio_completo, ctx_auth)
        carpeta = ctx.web.get_folder_by_server_relative_url(os.getenv("SHAREPOINT_DOC_PATH"))
        archivos = carpeta.files
        ctx.load(archivos)
        ctx.execute_query()

        return [archivo.properties["Name"] for archivo in archivos if archivo.properties["Name"].endswith('.xlsx')]
    
    except ClientRequestException as e:
        print(f"Error al listar archivos: {e}")
        return []

def descargar_excel_desde_sharepoint(nombre_archivo):
    try:
        sitio_url = os.getenv("SHAREPOINT_URL")
        sitio_completo = f"{sitio_url}/{os.getenv('SHAREPOINT_SITE')}"
        ctx_auth = AuthenticationContext(sitio_completo)
        if not ctx_auth.acquire_token_for_user(os.getenv("SHAREPOINT_USER"), os.getenv("SHAREPOINT_PASSWORD")):
            raise ValueError("Error de autenticación")
        
        ctx = ClientContext(sitio_completo, ctx_auth)
        file_url = f"{os.getenv('SHAREPOINT_DOC_PATH')}/{nombre_archivo}"
        file = ctx.web.get_file_by_server_relative_path(file_url)
        ctx.load(file)
        ctx.execute_query()

        file_content = BytesIO()
        file.download(file_content).execute_query()
        file_content.seek(0)
        return file_content
    except Exception as e:
        print(f"Error al descargar {nombre_archivo}: {e}")
        return None

def procesar_todos_los_excel():
    archivos_excel = listar_archivos_excel_en_sharepoint()
    indice_general = []
    
    for archivo in archivos_excel:
        print(f"Procesando: {archivo}")
        contenido = descargar_excel_desde_sharepoint(archivo)
        if contenido:
            nombre_base = os.path.splitext(archivo)[0]
            print(f"Creando carpeta para: {nombre_base}")
            carpeta_salida = f"html_output/{nombre_base}"
            os.makedirs(carpeta_salida, exist_ok=True)
            
            # Procesar el archivo (usando tu función existente)
            indice_archivo = excel_a_html_multiple(nombre_base,contenido, carpeta_salida)
            carpeta_salida =f"{nombre_base}"## Uso unico para el nombre de la carpeta
            indice_general.append({
                'nombre': nombre_base,
                'archivo': f"{carpeta_salida}/index.html"
            })
    
    # Generar índice general
    if indice_general:
        generar_indice_general(indice_general)
        print(f"Índice general creado en: html_output/indice.html")
    else:
        print("No se encontraron archivos .xlsx en la carpeta.")
    
    return indice_general

def es_texto_aislado(ws, row_idx, merged_cells):
    """
    Determina si una fila contiene texto aislado (no parte de una tabla estructurada).
    """
    row = ws[row_idx]
    # Contar celdas con contenido
    celdas_con_contenido = sum(1 for cell in row if cell.value is not None)
    # Si hay menos de 2 celdas con contenido, es texto aislado
    if celdas_con_contenido < 2:
        return True
    # Verificar si es un título de sección
    first_cell_value = str(row[0].value).strip() if row[0].value else ""
    if first_cell_value and first_cell_value.isupper() and len(first_cell_value) > 10:
        return True
    return False

def procesar_texto_aislado(ws, row_idx, merged_cells, sheet_name):
    """
    Procesa una fila de texto aislado y devuelve el HTML correspondiente.
    """
    row = ws[row_idx]
    contenido = []
    for cell in row:
        cell_value = cell.value
        # Verificar si la celda es parte de un rango combinado
        for merged in merged_cells:
            if cell.coordinate in merged['range']:
                if (cell.row, cell.column) == merged['first_cell']:
                    cell_value = merged['value']
                else:
                    cell_value = None
                break
        if cell_value is not None:
            contenido.append(str(cell_value).strip())
    # No agregar nada si el texto es igual al nombre de la hoja (sheet_name)
    if contenido and " ".join(contenido).strip() == sheet_name:
        return ""
    texto = " ".join(contenido).strip()
    # Determinar si es un título
    if texto.isupper() and len(texto) > 7:
        return f'<h2 class="titulo-seccion">{texto}</h2>\n'
    else:
        return f'<div class="texto-contenido">{texto}</div>\n'

def procesar_tabla(ws, start_row, merged_cells):
    """
    Procesa una tabla desde la fila start_row y devuelve el HTML y número de filas procesadas.
    Combina celdas de encabezado vacías adyacentes usando colspan.
    """
    # Determinar el alcance de la tabla
    end_row = start_row
    while end_row <= ws.max_row:
        if es_texto_aislado(ws, end_row, merged_cells):
            break
        end_row += 1
    # Crear DataFrame con el rango de la tabla
    data = []
    for row_idx in range(start_row, end_row):
        row_data = []
        for cell in ws[row_idx]:
            cell_value = cell.value
            cell_number_format = cell.number_format # Obtener el formato de número de la celda
            # Manejar celdas combinadas
            for merged in merged_cells:
                if cell.coordinate in merged['range']:
                    if (cell.row, cell.column) == merged['first_cell']:
                        cell_value = merged['value']
                    else:
                        cell_value = None
                    break
            # Process percentage values
            if isinstance(cell_value, (int, float)) and '%' in cell_number_format:
                cell_value = f"{cell_value:.2%}"# Convertir a porcentaje con dos decimales
            elif cell_value is not None:
                cell_value = str(cell_value).strip()# Convertir a cadena
            row_data.append(cell_value)
        data.append(row_data)

    # Limpiar DataFrame (eliminar filas/columnas completamente vacías)
    df = pd.DataFrame(data).dropna(how='all').dropna(axis=1, how='all')
    # Generar HTML de la tabla
    html = '<div class="tabla-contenedor">\n'
    if len(df) > 0:
        # Procesar encabezados para combinar celdas vacías adyacentes
        headers = df.iloc[0].tolist()
        processed_headers = []
        i = 0
        while i < len(headers):
            if headers[i] is not None and str(headers[i]).strip() != '':
                # Contar celdas vacías siguientes
                colspan = 1
                j = i + 1
                while j < len(headers) and (headers[j] is None or str(headers[j]).strip() == ''):
                    colspan += 1
                    j += 1
                if colspan > 1:
                    processed_headers.append(f'<th colspan="{colspan}">{headers[i]}</th>')
                else:
                    processed_headers.append(f'<th>{headers[i]}</th>')
                i = j
            else:
                i += 1
        # Crear la tabla HTML con encabezados procesados
        html += '<table class="tabla-estructurada">\n'
        html += '  <thead>\n    <tr>\n'
        html += '      ' + '\n      '.join(processed_headers) + '\n'
        html += '    </tr>\n  </thead>\n'
        # Procesar el cuerpo de la tabla
        html += '  <tbody>\n'
        for _, row in df.iloc[1:].iterrows():
            html += '    <tr>\n'
            for cell_value in row:
                if isinstance(cell_value, str) and cell_value.endswith('%'):
                    html += f'      <td class="percentage-cell">{cell_value}</td>\n'
                else:
                    html += f'      <td>{cell_value if cell_value is not None else ""}</td>\n'
            html += '    </tr>\n'
        html += '  </tbody>\n</table>\n'
    else:
        html += '<p>Tabla vacía</p>\n'
    html += '</div>\n'
    return html, end_row - start_row

def slugify(texto):
    texto = re.sub(r'[^\w\s-]', '', texto.lower())
    return re.sub(r'[-\s]+', '-', texto).strip('-_')

def generar_html_hoja(ws, sheet_name, nombre_archivo_excel):
    """Genera el HTML para una hoja específica"""
    # Procesar celdas combinadas
    merged_cells = []
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = range_boundaries(str(merged_range))
        merged_cells.append({
            'range': merged_range,
            'value': ws.cell(row=min_row, column=min_col).value,
            'first_cell': (min_row, min_col)
        })
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset=\"UTF-8\">
    <title>{sheet_name} - {nombre_archivo_excel}</title>
    <link rel="stylesheet" href="../css/styles.css">
    <link rel=\"stylesheet\" href=\"https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.3/css/all.min.css\">
</head>
<body>
    <header>
        <h1>{sheet_name}</h1>
        <a href=\"index.html\" class=\"btn-volver\"><i class=\"fas fa-arrow-left\"></i> Volver al índice</a>
    </header>
    <div class=\"contenido-hoja\">
"""
    #Procesar filas de la hoja
    current_row = 1
    while current_row <= ws.max_row:
        # Verificar si la fila está vacía
        row_empty = all(cell.value is None for cell in ws[current_row])
        if not row_empty:
            if es_texto_aislado(ws, current_row, merged_cells):
                html += procesar_texto_aislado(ws, current_row, merged_cells, sheet_name)
                current_row += 1
            else:
                # Procesar tabla
                table_html, rows_processed = procesar_tabla(ws, current_row, merged_cells)
                html += table_html
                current_row += rows_processed
        else:
            current_row += 1
    html += f"""
    </div>
    <footer>
        <div class=\"footer-flex\">
            <img src="../assets/azul_sdgdtic.png" alt="Logo secretaria de Gobierno Digital y Tecnología de la Información y Comunicaciones">
            <span>&copy; 2025</span>
        </div>
    </footer>
</body>
</html>
"""
    return html

def generar_indice(indice, carpeta_salida, nombre_archivo_excel):
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset=\"UTF-8\">
    <title>Indice - {nombre_archivo_excel}</title>
    <link rel="stylesheet" href="../css/styles.css">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.3/css/all.min.css">
</head>
<body>
    <header><h1>{nombre_archivo_excel}
    </h1>
    <a href="../indice.html " class="btn-volver"><i class="fas fa-arrow-left"></i> Volver al índice de Libros</a>
    </header>
    <div class=\"contenedor-indice\">
        <h2>Indice de Contenidos</h2>
        <ul class=\"lista-indice\">
"""
    for item in indice:
        html += f'<li><a href="{item["archivo"]}"><i class="fas fa-file-alt"></i> {item["nombre"]}</a></li>\n'
   
    html += f"""        </ul>
        </div>
        <footer>
            <div class="footer-flex">
                <img src="../assets/azul_sdgdtic.png" alt="Logo secretaria de Gobierno Digital y Tecnología de la Información y Comunicaciones">
                <span>&copy; 2025</span>
            </div>
        </footer>
    </body>
    </html>
    """
    # Guardar archivos
    with open(os.path.join(carpeta_salida, 'index.html'), 'w', encoding='utf-8') as f:
        f.write(html)

def excel_a_html_multiple(nombre_base, contenido_excel, carpeta_salida='html_output'):  # <-- Ahora recibe 2 parámetros
    if contenido_excel is None:
        return []
    
    wb = load_workbook(contenido_excel, data_only=True)
    os.makedirs(carpeta_salida, exist_ok=True)
    
    # Obtener el nombre del archivo desde .env
    nombre_archivo = nombre_base
    if not nombre_archivo:
        raise ValueError("La variable SHAREPOINT_FILE no está definida en .env")
    
    nombre_archivo_excel = os.path.splitext(nombre_archivo)[0]
    print(f"Procesando archivo: {nombre_archivo_excel}")
    
    indice = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in ["índice", "datoscbox"]:
            continue
        
        ws = wb[sheet_name]
        titulo_hoja = None
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    titulo_hoja = str(cell.value).strip()
                    break
            if titulo_hoja is not None:
                break
        
        nombre_archivo_html = f"{slugify(sheet_name)}.html"
        indice.append({'nombre': titulo_hoja, 'archivo': nombre_archivo_html})
        
        html = generar_html_hoja(ws, titulo_hoja, nombre_archivo_excel)
        
        with open(os.path.join(carpeta_salida, nombre_archivo_html), 'w', encoding='utf-8') as f:
            f.write(html)
    
    generar_indice(indice, carpeta_salida, nombre_archivo_excel)
    return indice


def generar_indice_general(indice, carpeta_salida='html_output'):

    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset=\"UTF-8\">
    <title>Libros de Dominios</title>
    <link rel=\"stylesheet\" href=\"css/styles.css\">
</head>
<body>
    <header><h1>Libros de Dominos</h1></header>
    <div class=\"contenedor-indice\">
        <h2>Indice de Contenidos</h2>
        <ul class=\"lista-indice\">
"""
    for item in indice:
        html += f'<li><a href="{item["archivo"]}">{item["nombre"]}</a></li>\n'
    
    html += f"""        </ul>
        </div>
        <footer>
            <div class="footer-flex">
                <img src="assets/azul_sdgdtic.png" alt="Logo secretaria de Gobierno Digital y Tecnología de la Información y Comunicaciones">
                <span>&copy; 2025</span>
            </div>
        </footer>
    </body>
    </html>
    """
    
    with open(f"{carpeta_salida}/indice.html", 'w', encoding='utf-8') as f:
        f.write(html)