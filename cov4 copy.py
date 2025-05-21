import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import os


def es_texto_aislado(ws, row_idx, merged_cells):
    """
    Determina si una fila contiene texto aislado (no parte de una tabla estructurada).
    """
    row = ws[row_idx]
    
    # Contar celdas con contenido
    celdas_con_contenido = sum(1 for cell in row if cell.value is not None)
    
    # Si hay menos de 2 celdas con contenido, es texto aislado
    if celdas_con_contenido < 2:
        return True
    
    # Verificar si es un título de sección
    first_cell_value = str(row[0].value).strip() if row[0].value else ""
    if first_cell_value and first_cell_value.isupper() and len(first_cell_value) > 10:
        return True
    
    return False

def procesar_texto_aislado(ws, row_idx, merged_cells):
    """
    Procesa una fila de texto aislado y devuelve el HTML correspondiente.
    """
    row = ws[row_idx]
    contenido = []
    
    for cell in row:
        cell_value = cell.value
        
        # Verificar si la celda es parte de un rango combinado
        for merged in merged_cells:
            if cell.coordinate in merged['range']:
                if (cell.row, cell.column) == merged['first_cell']:
                    cell_value = merged['value']
                else:
                    cell_value = None
                break
        
        if cell_value is not None:
            contenido.append(str(cell_value).strip())
    
    texto = " ".join(contenido).strip()
    
    # Determinar si es un título
    if texto.isupper() and len(texto) > 10:
        return f'<h2 class="titulo-seccion">{texto}</h2>\n'
    else:
        return f'<div class="texto-contenido">{texto}</div>\n'

def procesar_tabla(ws, start_row, merged_cells):
    """
    Procesa una tabla desde la fila start_row y devuelve el HTML y número de filas procesadas.
    Combina celdas de encabezado vacías adyacentes usando colspan.
    """
    # Determinar el alcance de la tabla
    end_row = start_row
    while end_row <= ws.max_row:
        if es_texto_aislado(ws, end_row, merged_cells):
            break
        end_row += 1
    
    # Crear DataFrame con el rango de la tabla
    data = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row-1):
        row_data = []
        for cell in row:
            cell_value = cell.value
            
            # Manejar celdas combinadas
            for merged in merged_cells:
                if cell.coordinate in merged['range']:
                    if (cell.row, cell.column) == merged['first_cell']:
                        cell_value = merged['value']
                    else:
                        cell_value = None
                    break
            
            row_data.append(cell_value)
        data.append(row_data)
    
    df = pd.DataFrame(data)
    
    # Limpiar DataFrame (eliminar filas/columnas completamente vacías)
    df = df.dropna(how='all').dropna(axis=1, how='all')
    
    # Generar HTML de la tabla
    html = '<div class="tabla-contenedor">\n'
    
    if len(df) > 0:
        # Procesar encabezados para combinar celdas vacías adyacentes
        headers = df.iloc[0].tolist()
        processed_headers = []
        i = 0
        while i < len(headers):
            if headers[i] is not None and str(headers[i]).strip() != '':
                # Contar celdas vacías siguientes
                colspan = 1
                j = i + 1
                while j < len(headers) and (headers[j] is None or str(headers[j]).strip() == ''):
                    colspan += 1
                    j += 1
                
                if colspan > 1:
                    processed_headers.append(f'<th colspan="{colspan}">{headers[i]}</th>')
                else:
                    processed_headers.append(f'<th>{headers[i]}</th>')
                i = j
            else:
                i += 1
        
        # Construir manualmente la tabla HTML
        html += '<table class="tabla-estructurada">\n'
        html += '  <thead>\n    <tr>\n'
        html += '      ' + '\n      '.join(processed_headers) + '\n'
        html += '    </tr>\n  </thead>\n'
        
        # Procesar el cuerpo de la tabla
        html += '  <tbody>\n'
        for _, row in df.iloc[1:].iterrows():
            html += '    <tr>\n'
            for cell in row:
                html += f'      <td>{cell if cell is not None else ""}</td>\n'
            html += '    </tr>\n'
        html += '  </tbody>\n</table>\n'
    else:
        html += '<p>Tabla vacía</p>\n'
    
    html += '</div>\n'
    
    return html, end_row - start_row

def deberia_ser_encabezado(df):
    """
    Determina si la primera fila del DataFrame debería ser tratada como encabezado.
    """
    
    
    primera_fila = df.iloc[0]
    resto_tabla = df.iloc[1:]
    
    # Contar valores de texto en la primera fila vs el resto
    textos_primera_fila = sum(isinstance(x, str) for x in primera_fila)
    textos_resto = sum(isinstance(x, str) for x in resto_tabla.values.flatten())
    
    return textos_primera_fila > textos_resto / max(1, len(resto_tabla))



def excel_a_html_multiple(archivo_excel, carpeta_salida='html_output'):
    # Crear carpeta de salida si no existe
    os.makedirs(carpeta_salida, exist_ok=True)
    
    # Cargar el libro de Excel
    wb = load_workbook(archivo_excel, data_only=True)
    
    # Lista para almacenar información del índice
    indice = []
    
    # Obtener el nombre del archivo Excel sin la extensión
    nombre_archivo_excel = os.path.splitext(os.path.basename(archivo_excel))[0]
    print(f"Procesando archivo: {nombre_archivo_excel}")
    
    # Procesar cada hoja del libro
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        nombre_archivo = f"{slugify(sheet_name)}.html"
        indice.append({'nombre': sheet_name, 'archivo': nombre_archivo})
        
        # Crear HTML para esta hoja
        html = generar_html_hoja(ws, sheet_name, nombre_archivo_excel)
        
        # Guardar archivo HTML
        with open(os.path.join(carpeta_salida, nombre_archivo), 'w', encoding='utf-8') as f:
            f.write(html)
    
    # Generar archivo índice
    generar_indice(indice, carpeta_salida, nombre_archivo_excel)
    
    return indice

def slugify(texto):
    """Convierte un texto a formato válido para nombre de archivo"""
    import re
    texto = re.sub(r'[^\w\s-]', '', texto.lower())
    return re.sub(r'[-\s]+', '-', texto).strip('-_')

def generar_html_hoja(ws, sheet_name, nombre_archivo_excel):
    

    """Genera el HTML para una hoja específica"""
    # Procesar celdas combinadas
    merged_cells = []
    for merged_range in ws.merged_cells.ranges if ws.merged_cells else []:
        min_row, min_col, max_row, max_col = range_boundaries(str(merged_range))
        merged_cells.append({
            'range': merged_range,
            'value': ws.cell(row=min_row, column=min_col).value,
            'first_cell': (min_row, min_col)
        })
    
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{sheet_name} - {nombre_archivo_excel}</title>
    <link rel="stylesheet" href="styles.css">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.3/css/all.min.css">
</head>
<body>
    <header>
        <h1>{sheet_name}</h1>
        <a href="index.html" class="btn-volver"><i class="fas fa-arrow-left"></i> Volver al índice</a>
    </header>
    <div class="contenido-hoja">
"""
    
    # Procesar filas en orden
    current_row = 1
    while current_row <= ws.max_row:
        # Verificar si la fila está vacía
        row_empty = all(cell.value is None for cell in ws[current_row])
        
        if not row_empty:
            # Determinar si es texto aislado o parte de una tabla
            if es_texto_aislado(ws, current_row, merged_cells):
                html += procesar_texto_aislado(ws, current_row, merged_cells)
                current_row += 1
            else:
                # Procesar como tabla
                table_html, rows_processed = procesar_tabla(ws, current_row, merged_cells)
                html += table_html
                current_row += rows_processed
        else:
            current_row += 1
    
    html += f"""
    </div>
    <footer>
        <p>{nombre_archivo_excel} - &copy; 2025</p>
    </footer>
</body>
</html>
"""
    return html

def generar_indice(indice, carpeta_salida, nombre_archivo_excel):
    """Genera el archivo índice HTML"""
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Índice - {nombre_archivo_excel}</title>
    <link rel="stylesheet" href="styles.css">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.3/css/all.min.css">
</head>
<body>
    <header>
        <h1>{nombre_archivo_excel}</h1>
    </header>
    <div class="contenedor-indice">
        <h2>Índice de Contenidos</h2>
        <ul class="lista-indice">
"""
    
    for item in indice:
        html += f'            <li><a href="{item["archivo"]}"><i class="fas fa-file-alt"></i> {item["nombre"]}</a></li>\n'
    
    html += f"""        </ul>
    </div>
    <footer>
        <p>{nombre_archivo_excel} - &copy; 2077</p>
    </footer>
</body>
</html>
"""
    
    # Generar archivo CSS
    css = """/* Estilos generales */
body {
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    line-height: 1.6;
    margin: 0;
    padding: 0;
    color: #333;
    background-color: #f5f7fa;
}

header {
    background-color: #2c3e50;
    color: white;
    padding: 20px;
    text-align: center;
    box-shadow: 0 2px 5px rgba(0,0,0,0.1);
}

h1, h2, h3 {
    margin: 0;
    font-weight: 600;
}

.contenido-hoja, .contenedor-indice {
    max-width: 1200px;
    margin: 30px auto;
    padding: 0 20px;
}

/* Estilos para el índice */
.lista-indice {
    list-style-type: none;
    padding: 0;
    margin: 30px 0;
}

.lista-indice li {
    margin-bottom: 10px;
    background-color: white;
    border-radius: 5px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    transition: transform 0.2s;
}

.lista-indice li:hover {
    transform: translateX(5px);
}

.lista-indice a {
    display: block;
    padding: 15px 20px;
    color: #2c3e50;
    text-decoration: none;
    font-size: 18px;
}

.lista-indice a:hover {
    color: #3498db;
}

.lista-indice i {
    margin-right: 10px;
    color: #3498db;
}

/* Estilos para el contenido de las hojas */
.texto-contenido {
    margin: 20px 0;
    padding: 15px;
    background-color: white;
    border-radius: 5px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    line-height: 1.8;
}

.titulo-seccion {
    font-size: 22px;
    color: #2c3e50;
    margin: 40px 0 20px 0;
    padding-bottom: 10px;
    border-bottom: 2px solid #3498db;
}

.tabla-contenedor {
    margin: 30px 0;
    overflow-x: auto;
    background-color: white;
    border-radius: 5px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    padding: 1px;
}

table {
    border-collapse: collapse;
    width: 100%;
    margin: 0;
}

th {
    background-color: #3498db;
    color: white;
    font-weight: 600;
    padding: 12px 15px;
    text-align: center;
}

td {
    padding: 10px 15px;
    border: 1px solid #e0e0e0;
    vertical-align: top;
}

tr:nth-child(even) {
    background-color: #f8f9fa;
}

.btn-volver {
    display: inline-block;
    margin-top: 20px;
    padding: 10px 15px;
    background-color: #3498db;
    color: white;
    text-decoration: none;
    border-radius: 5px;
    transition: background-color 0.3s;
}

.btn-volver:hover {
    background-color: #2980b9;
}

.btn-volver i {
    margin-right: 5px;
}

footer {
    text-align: center;
    padding: 20px;
    background-color: #2c3e50;
    color: white;
    margin-top: 40px;
}
"""
    
    # Guardar archivos
    with open(os.path.join(carpeta_salida, 'index.html'), 'w', encoding='utf-8') as f:
        f.write(html)
    
    with open(os.path.join(carpeta_salida, 'styles.css'), 'w', encoding='utf-8') as f:
        f.write(css)
    
    return html

## Dar formato a la salida


if __name__ == "__main__":
    print("Procesando archivo Excel...")
    indice = excel_a_html_multiple('test01.xlsx')
    print(f"Proceso completado. Se generaron {len(indice)} archivos HTML.")
    print(f"Archivo índice creado en: html_output/index.html")
    